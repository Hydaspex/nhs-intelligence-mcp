"""Ingest NHS England's monthly "Incomplete Provider" RTT XLSX into the CSV
contract that :class:`nhs_intel.sources.rtt.RttCsvSource` consumes.

NHS England publishes one workbook per month with a ``Provider`` sheet keyed
on ``(Provider Code, Treatment Function)``. Each row already carries a column
named ``Average (median) waiting time (in weeks)``, so this ingest step reads
that figure directly rather than deriving a summary statistic from the
week-band distribution columns (``>0-1``, ``>1-2``, ... ``104 plus``) that
precede it in the sheet. The median is NHS England's own published summary
figure for the wait, so using it keeps this ingest a straight passthrough
instead of introducing a second, home-grown statistic.

Rows are skipped in two cases, both counted and reported on stdout:
  * the median cell holds NHS England's suppression marker (a literal ``"-"``
    string, used for small counts) rather than a number.
  * the treatment function code is ``C_999`` ("Total"), which is a
    per-provider aggregate across all specialties, not a specialty row, and
    would double-count against the per-specialty rows if kept.

HTTP is isolated behind a small ``TextHttpClient`` Protocol (the same seam
used by ``sources/cqc.py``), so the index-page-parsing function is unit
tested with a fixture HTML string and no network is touched in the offline
suite. The XLSX-parsing function takes a local path and is tested against a
small fixture workbook, also with no network.
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
import urllib.error
import urllib.request
from collections.abc import Iterator
from datetime import date
from pathlib import Path
from typing import NamedTuple, Protocol, runtime_checkable

from openpyxl import load_workbook

_SHEET_NAME = "Provider"
_HEADER_ROW = 14  # 1-indexed; row 13 is a section label, row 14 has headers
_TOTAL_TREATMENT_FUNCTION_CODE = "C_999"
_SUPPRESSED_MARKER = "-"
_CSV_FIELDNAMES = ["provider_code", "specialty", "weeks", "as_of"]


class RttRow(NamedTuple):
    """One provider/specialty wait figure, matching the RttCsvSource contract."""

    provider_code: str
    specialty: str
    weeks: float
    as_of: date


@runtime_checkable
class TextHttpClient(Protocol):
    """Minimal HTTP GET seam returning response body text, for HTML pages."""

    def get_text(self, url: str) -> str:
        """GET a URL and return the decoded body text."""
        ...


class UrllibTextHttpClient:
    """The real HTTP client: a thin urllib GET returning decoded text."""

    def __init__(self, timeout: float = 30.0) -> None:
        self._timeout = timeout

    def get_text(self, url: str) -> str:
        request = urllib.request.Request(url, headers={"Accept": "text/html"})
        with urllib.request.urlopen(request, timeout=self._timeout) as response:
            return response.read().decode("utf-8")


def find_incomplete_provider_url(index_html: str, month_label: str) -> str:
    """Find the "Incomplete Provider <month_label>" XLSX link in an index page.

    ``month_label`` is NHS England's short form, e.g. ``"Jun26"``. The index
    page lists several XLSX files per month (Admitted, NonAdmitted,
    Incomplete, each split Provider/Commissioner); this matches on the link
    text "Incomplete Provider <month_label>" to pick the right one, since the
    download URL itself carries an unpredictable trailing upload code and
    cannot be constructed from a pattern.

    Raises:
        LookupError: if no matching link is found for the given month.
    """
    pattern = re.compile(
        rf'<a [^>]*href="([^"]+\.xlsx)"[^>]*>\s*Incomplete Provider {re.escape(month_label)}\b',
        re.IGNORECASE,
    )
    match = pattern.search(index_html)
    if not match:
        raise LookupError(f"No 'Incomplete Provider {month_label}' link found in index page")
    return match.group(1)


def resolve_incomplete_provider_url(
    index_url: str, month_label: str, http: TextHttpClient | None = None
) -> str:
    """Fetch the index page and resolve the XLSX URL for one month."""
    client = http or UrllibTextHttpClient()
    html = client.get_text(index_url)
    return find_incomplete_provider_url(html, month_label)


def download_xlsx(url: str, dest_path: Path, timeout: float = 60.0) -> None:
    """Download the XLSX at ``url`` to ``dest_path``, overwriting if present."""
    request = urllib.request.Request(url, headers={"Accept": "application/octet-stream"})
    try:
        with urllib.request.urlopen(request, timeout=timeout) as response:
            dest_path.write_bytes(response.read())
    except urllib.error.HTTPError as exc:
        raise LookupError(f"Failed to download RTT XLSX from {url}: {exc}") from exc


def parse_incomplete_provider_xlsx(xlsx_path: Path, as_of: date) -> Iterator[RttRow]:
    """Yield one RttRow per provider/specialty in the "Provider" sheet.

    Skips the per-provider "Total" aggregate row (treatment function code
    ``C_999``) and any row whose median wait is suppressed (NHS England marks
    small counts with ``"-"`` instead of a number).
    """
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet = workbook[_SHEET_NAME]
    rows = sheet.iter_rows(min_row=_HEADER_ROW + 1, values_only=True)
    header = next(sheet.iter_rows(min_row=_HEADER_ROW, max_row=_HEADER_ROW, values_only=True))
    columns = _resolve_columns(header)

    for row in rows:
        parsed = _parse_data_row(row, columns, as_of)
        if parsed is not None:
            yield parsed


class _Columns(NamedTuple):
    provider_code: int
    treatment_function_code: int
    treatment_function: int
    median_weeks: int


def _resolve_columns(header: tuple) -> _Columns:
    """Map the header row's expected column names to their positions."""
    index_by_name = {name: i for i, name in enumerate(header) if name is not None}
    required = {
        "provider_code": "Provider Code",
        "treatment_function_code": "Treatment Function Code",
        "treatment_function": "Treatment Function",
        "median_weeks": "Average (median) waiting time (in weeks)",
    }
    missing = [label for label in required.values() if label not in index_by_name]
    if missing:
        raise ValueError(f"Incomplete Provider XLSX missing expected columns: {missing}")
    return _Columns(**{key: index_by_name[label] for key, label in required.items()})


def _parse_data_row(row: tuple, columns: _Columns, as_of: date) -> RttRow | None:
    """Build one RttRow from a raw sheet row, or None if it should be skipped."""
    provider_code = row[columns.provider_code]
    if provider_code is None:
        return None
    if row[columns.treatment_function_code] == _TOTAL_TREATMENT_FUNCTION_CODE:
        return None

    median = row[columns.median_weeks]
    if isinstance(median, str) or median is None:
        return None  # suppressed (NHS England marks small counts with "-")

    return RttRow(
        provider_code=str(provider_code).strip(),
        specialty=str(row[columns.treatment_function]).strip(),
        weeks=float(median),
        as_of=as_of,
    )


def write_csv(rows: Iterator[RttRow], out_path: Path) -> int:
    """Write rows to ``out_path`` in the RttCsvSource CSV contract. Returns count."""
    count = 0
    with out_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(_CSV_FIELDNAMES)
        for row in rows:
            writer.writerow([row.provider_code, row.specialty, row.weeks, row.as_of.isoformat()])
            count += 1
    return count


def month_label_and_as_of(year_month: str) -> tuple[str, date]:
    """Convert "2026-06" into NHS England's short label "Jun26" and the as_of date."""
    parsed = date.fromisoformat(f"{year_month}-01")
    label = parsed.strftime("%b%y")
    return label, parsed


def _build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--index-url",
        required=True,
        help="RTT data index page for the target NHS financial year, e.g. "
        "https://www.england.nhs.uk/.../rtt-data-2026-27/",
    )
    parser.add_argument(
        "--month",
        required=True,
        help="Target reporting month as YYYY-MM, e.g. 2026-06",
    )
    parser.add_argument("--out", required=True, type=Path, help="Output CSV path")
    parser.add_argument(
        "--xlsx-cache",
        type=Path,
        default=None,
        help="Where to save the downloaded XLSX (default: alongside --out)",
    )
    return parser


def run(index_url: str, year_month: str, out_path: Path, xlsx_cache: Path | None) -> int:
    """Fetch, download, parse and write one month's RTT ingest. Returns row count."""
    month_label, as_of = month_label_and_as_of(year_month)
    xlsx_path = xlsx_cache or out_path.with_suffix(".xlsx")

    print(f"Resolving 'Incomplete Provider {month_label}' link from {index_url}")
    url = resolve_incomplete_provider_url(index_url, month_label)
    print(f"Found: {url}")

    print(f"Downloading to {xlsx_path}")
    download_xlsx(url, xlsx_path)

    print(f"Parsing {xlsx_path} (as_of={as_of.isoformat()})")
    rows = list(parse_incomplete_provider_xlsx(xlsx_path, as_of))

    print(f"Writing {len(rows)} rows to {out_path}")
    written = write_csv(iter(rows), out_path)
    print(f"Done: {written} rows written")
    return written


def main() -> None:
    args = _build_arg_parser().parse_args()
    try:
        run(args.index_url, args.month, args.out, args.xlsx_cache)
    except (LookupError, ValueError, OSError) as exc:
        print(f"RTT ingest failed: {exc}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
